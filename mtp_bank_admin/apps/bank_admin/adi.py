from collections import defaultdict
from datetime import date
from decimal import Decimal
import logging

from django.conf import settings
from openpyxl import load_workbook, styles
from openpyxl.writer.excel import save_virtual_workbook

from . import ADI_JOURNAL_LABEL
from . import adi_config as config
from .types import PaymentType, RecordType
from .exceptions import EmptyFileError
from .utils import (
    retrieve_all_transactions, retrieve_all_valid_credits,
    reconcile_for_date, retrieve_prisons, get_full_narrative
)

logger = logging.getLogger('mtp')


class AdiJournal(object):

    STYLE_TYPES = {
        'fill': styles.PatternFill,
        'border': styles.Border,
        'font': styles.Font,
        'alignment': styles.Alignment
    }

    def __init__(self, *args, **kwargs):
        self.wb = load_workbook(settings.ADI_TEMPLATE_FILEPATH, keep_vba=True)
        self.journal_ws = self.wb.get_sheet_by_name(config.ADI_JOURNAL_SHEET)

        self.current_row = config.ADI_JOURNAL_START_ROW

    def _next_row(self, increment=1):
        self.current_row += increment

    def _get_cell(self, field):
        return '%s%s' % (config.ADI_JOURNAL_FIELDS[field]['column'],
                         self.current_row)

    def _set_field(self, field, value, style=None, extra_style=None):
        cell = self._get_cell(field)
        self.journal_ws[cell] = value

        computed_style = defaultdict(dict)
        base_style = style or config.ADI_JOURNAL_FIELDS[field]['style']
        for key in base_style:
            computed_style[key].update(base_style[key])

        if extra_style:
            for key in extra_style:
                computed_style[key].update(extra_style[key])

        for key in computed_style:
            setattr(
                self.journal_ws[cell],
                key,
                self.STYLE_TYPES[key](**computed_style[key])
            )
        return self.journal_ws[cell]

    def _add_column_sum(self, field):
        self._set_field(
            field,
            ('=SUM(%(column)s%(start)s:%(column)s%(end)s)'
                % {
                    'column': config.ADI_JOURNAL_FIELDS[field]['column'],
                    'start': config.ADI_JOURNAL_START_ROW,
                    'end': self.current_row - 1
                }),
            extra_style=config.ADI_FINAL_ROW_STYLE
        )
        cell = self._get_cell(field)
        self.journal_ws[cell].number_format = '£#,##0.00_-'

    def _lookup(self, field, payment_type, record_type, context={}):
        try:
            value_dict = config.ADI_JOURNAL_FIELDS[field]['value']
            value = value_dict[payment_type.name][record_type.name]
            if value:
                return value.format(**context)
        except KeyError:
            pass  # no static value
        return None

    def _finish_journal(self, receipt_date, user):
        for field in config.ADI_JOURNAL_FIELDS:
            self._set_field(field, '', extra_style=config.ADI_FINAL_ROW_STYLE)
        bold = {'font': {'name': 'Arial', 'bold': True},
                'alignment': {'horizontal': 'left'}}

        self._set_field('upload', 'Totals:', extra_style=dict(config.ADI_FINAL_ROW_STYLE, **bold))
        self._add_column_sum('debit')
        self._add_column_sum('credit')

        self.wb.get_named_range('BNE_UPLOAD').destinations = [(
            self.journal_ws,
            "$B$%(start)s:$B$%(end)s" % {
                'start': config.ADI_JOURNAL_START_ROW,
                'end': self.current_row - 1,
            }
        )]

        self._next_row(increment=2)
        self._set_field('description', 'Uploaded by:', style=config._tan_style, extra_style=bold)

        self._next_row(increment=2)
        self._set_field('description', 'Checked by:', style=config._tan_style, extra_style=bold)

        self._next_row(increment=2)
        self._set_field('description', 'Posted by:', style=config._tan_style, extra_style=bold)

        batch_date = date.today().strftime(config.ADI_BATCH_DATE_FORMAT)
        self.journal_ws[config.ADI_BATCH_NAME_CELL] = config.ADI_BATCH_NAME_FORMAT % {
            'date': batch_date,
            'initials': user.get_initials() or '<initials>',
        }
        accounting_date = date.today()
        if accounting_date.month != receipt_date.month:
            accounting_date = receipt_date
        self.journal_ws[config.ADI_DATE_CELL] = accounting_date.strftime(config.ADI_DATE_FORMAT)
        self.journal_ws.title = receipt_date.strftime('%d%m%y')

    def add_payment_row(self, amount, payment_type, record_type, **kwargs):
        for field in config.ADI_JOURNAL_FIELDS:
            static_value = self._lookup(field, payment_type, record_type, context=kwargs)
            self._set_field(field, static_value)

        if record_type == RecordType.debit:
            self._set_field('debit', float(amount))
        elif record_type == RecordType.credit:
            self._set_field('credit', float(amount))

        self._next_row()

    def create_file(self, receipt_date, user):
        self._finish_journal(receipt_date, user)
        return (date.today().strftime(settings.ADI_OUTPUT_FILENAME),
                save_virtual_workbook(self.wb))


def generate_adi_journal(request, receipt_date):
    start_date, end_date = reconcile_for_date(request, receipt_date)

    credits = retrieve_all_valid_credits(
        request,
        received_at__gte=start_date,
        received_at__lt=end_date
    )
    refundable_transactions = retrieve_all_transactions(
        request,
        status='refundable',
        received_at__gte=start_date,
        received_at__lt=end_date
    )
    rejected_transactions = retrieve_all_transactions(
        request,
        status='unidentified',
        received_at__gte=start_date,
        received_at__lt=end_date
    )

    if (len(credits) == 0 and
            len(rejected_transactions) == 0 and
            len(refundable_transactions) == 0):
        raise EmptyFileError()

    journal_date = receipt_date.strftime('%d/%m/%Y')
    journal = AdiJournal()

    prisons = retrieve_prisons(request)
    bu_lookup = {prison['general_ledger_code']: nomis_id for nomis_id, prison in prisons.items()}

    debit_card_batches = defaultdict(int)
    prison_totals = defaultdict(int)
    prison_transactions = defaultdict(list)
    for credit in credits:
        business_unit = prisons[credit['prison']]['general_ledger_code']
        amount = Decimal(credit['amount'])/100
        prison_totals[business_unit] += amount
        if credit['source'] == 'online':
            if credit['reconciliation_code']:
                card_reconciliation_code = credit['reconciliation_code']
            else:
                card_reconciliation_code = '%s - Card payment' % journal_date
            debit_card_batches[card_reconciliation_code] += amount
        else:
            prison_transactions[business_unit].append(credit)

    # add valid payment rows
    # debit card batches
    for batch_code in debit_card_batches:
        journal.add_payment_row(
            debit_card_batches[batch_code], PaymentType.payment, RecordType.debit,
            reconciliation_code=batch_code
        )
    # other credits
    for business_unit in prison_totals:
        for transaction in prison_transactions.get(business_unit, []):
            journal.add_payment_row(
                Decimal(transaction['amount'])/100,
                PaymentType.payment, RecordType.debit,
                reconciliation_code=transaction['reconciliation_code']
            )
        journal.add_payment_row(
            prison_totals[business_unit], PaymentType.payment, RecordType.credit,
            prison_ledger_code=business_unit,
            prison_name=prisons[bu_lookup[business_unit]]['name'],
            date=journal_date
        )

    # add refund rows
    refund_total = 0
    for refund in refundable_transactions:
        refund_amount = Decimal(refund['amount'])/100
        refund_total += refund_amount
        journal.add_payment_row(
            refund_amount, PaymentType.refund, RecordType.debit,
            reconciliation_code=str(refund['ref_code'])
        )
    journal.add_payment_row(
        refund_total, PaymentType.refund, RecordType.credit, date=journal_date
    )

    # add reject rows
    for reject in rejected_transactions:
        reject_amount = Decimal(reject['amount'])/100
        reference = get_full_narrative(reject)
        journal.add_payment_row(
            reject_amount, PaymentType.reject, RecordType.debit,
            reconciliation_code=str(reject['ref_code'])
        )
        journal.add_payment_row(
            reject_amount, PaymentType.reject, RecordType.credit,
            reference=reference, date=journal_date
        )

    created_journal = journal.create_file(receipt_date, request.user)
    logger.info('{user} downloaded {label} containing {count} records'.format(
        user=request.user.username,
        label=ADI_JOURNAL_LABEL,
        count=len(credits) + len(rejected_transactions) + len(refundable_transactions)
    ))

    return created_journal
